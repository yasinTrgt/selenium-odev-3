from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl
from constants import globalConstants as c

class Test_KullaniciKilit:
     def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 
     
     def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()
     def getData():
        excel = openpyxl.load_workbook(c.INVALID_LOGIN_XLSX)
        sheet = excel["Sheet3"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        username = sheet.cell(row=2, column=1).value
        password = sheet.cell(row=2, column=2).value
        data.append((username,password))

        return data   
     @pytest.mark.parametrize("username,password",getData())
     def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        sleep(5)  # import time
        errorMessage = self.driver.find_element(By.XPATH,c.USER_BLOCK_XPATH)
        assert errorMessage.text == c.USER_BLOCK_DONT_MATCH
